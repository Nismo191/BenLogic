from datetime import datetime

from pandas.core.frame import DataFrame
from openpyxl import load_workbook
from numpy import NaN, nan

import file_operations
import flow
import REST_operations as rest
import pandas as pd

class Unit():
    def __init__(self, site, unit, scheme, tVal, tCharge, standCharge, serviceCharge,):
        self.site = site
        self.unit = unit
        self.scheme = scheme
        self.sum_transaction_value = tVal
        self.sum_transaction_charge = tCharge
        self.sum_standard_charge = standCharge
        self.sum_service_charge = serviceCharge
    


if __name__ == "__main__":
    data = file_operations.read_excel('Input/Upay example- 1.xlsx', 48)
    template = load_workbook('Input/1800 Upay Journal - Copy.xlsx')
    df_temp = []
    sites_array = ['Norton Canes', 'Stafford', 'Tibshelf', 'Tibshelf (North)', 'Tibshelf (South)']
    units_array = ['Costa Coffee', 'FFC', 'Restbite', 'WHSmiths', 'Chozen', 'Drive Thru', 'Leon']
    card_rates = {
        "Mastercard":{'rate':0.0108},
        "Visa":{'rate':0.0117},
        "Maestro":{'rate':0.0117},
        "American Express":{'rate':0},
        "Visa Debit": {'rate': 0.0117},
        "Mastercard Debit": {'rate': 0.0108},
        "Non-Standard (Credit)": {'rate': 0.0117},
        "Non-Standard (Debit)": {'rate': 0.0117},
        "Electron": {'rate': 0.0117},
        "Solo": {'rate': 0.0117},
        "JCB": {'rate': 0.0242},
        "Mastercard (Int'l)": {'rate': 0.0117},
        "Mastercard Debit (Int'l)": {'rate': 0.0117},
        "Visa (Europe)": {'rate': 0.0117},
        "Visa Debit (Europe)": {'rate': 0.0117},
        "Mastercard (Europe)": {'rate': 0.0117},
        "Visa (Int'l)": {'rate': 0.0117},
        "Visa Debit (Int'l)": {'rate': 0.0117},
        "Maestro (Europe)": {'rate': 0.0117},
        "Mastercard Debit (Europe)": {'rate': 0.0117},
        "Maestro (Int'l)": {'rate': 0.0117},
        "Mastercard (Corp/Prem)": {'rate': 0.0117},
        "Visa Debit (Corp/Prem)": {'rate': 0.0117},
        "Visa (Corp/Prem)": {'rate': 0.0117},
        "Mastercard Debit (Corp/Prem)": {'rate': 0.0117}
    }
    site = ""
    unit = ""
    rate = "0"
    sum_transaction_value = 0
    sum_transaction_charge = 0
    sum_standard_charge = 0
    sum_service_charge = 0
    totals = []

    # Remove pandas column
    for col in data.columns:
        if col.startswith('Unnamed'):
            data.drop(col, axis=1, inplace=True)

    for row in data.itertuples():
        # if row.Type contains 'Norton Canes'
        if any(x in str(row.Type) for x in sites_array):
            site = row.Type[5:]
        if any(x in str(row.Type) for x in units_array):
            unit = row.Type[6:]
        # Check if key is in card_rates
        if row._2 in card_rates:
            rate = card_rates[row._2]['rate']
        row = row + (site,unit,float(row.Count) * float(rate))
        df_temp.append(row)

    df = pd.DataFrame(df_temp)

    for row in df.itertuples():
        if pd.isna(row._3):
            df.drop(row.Index, axis=0, inplace=True)



    df.columns = ['Index', 'Type', 'Scheme', 'TCount', 'TVal', 'Standard Charge Rate', 'Standard Charge Value', 'Service Charge Rate', 'Service Charge Value', 'Site', 'Unit', 'Service Charge Value Corrected']

    for site in sites_array:
        for unit in units_array:
            sum_transaction_value = 0
            sum_transaction_charge = 0
            sum_standard_charge = 0
            sum_service_charge = 0
            if df.query('Site == "'+site+'" & Unit == "'+unit+'" & Scheme != "American Express"')['TVal'].sum() != 0:
                # print(site, unit, "Standard")
                # print(df.query('Site == "'+site+'" & Unit == "'+unit+'" & Scheme != "American Express"')['TVal'].sum())
                sum_transaction_value += df.query('Site == "'+site+'" & Unit == "'+unit+'" & Scheme != "American Express"')['TVal'].sum()
                sum_transaction_charge += (df.query('Site == "'+site+'" & Unit == "'+unit+'" & Scheme != "American Express"')['Service Charge Value Corrected'].sum())
                sum_transaction_charge += df.query('Site == "'+site+'" & Unit == "'+unit+'" & Scheme != "American Express"')['Standard Charge Value'].sum()
                sum_standard_charge += df.query('Site == "'+site+'" & Unit == "'+unit+'" & Scheme != "American Express"')['Service Charge Value'].sum()
                sum_service_charge = (sum_standard_charge/100)*20
                sum_transaction_value = float("{:.2f}".format(sum_transaction_value))
                sum_transaction_charge = float("{:.2f}".format(sum_transaction_charge))
                sum_standard_charge = float("{:.2f}".format(sum_standard_charge))
                sum_service_charge = float("{:.2f}".format(sum_service_charge))
                totals.append(Unit(site, unit, "", sum_transaction_value, sum_transaction_charge, sum_standard_charge, sum_service_charge))
            if df.query('Site == "'+site+'" & Unit == "'+unit+'" & Scheme == "American Express"')['TVal'].sum() != 0:
                sum_transaction_value = 0
                sum_transaction_charge = 0
                sum_standard_charge = 0
                sum_service_charge = 0
                # print(site, unit, "American Express")
                # print(df.query('Site == "'+site+'" & Unit == "'+unit+'" & Scheme == "American Express"')['TVal'].sum())
                sum_transaction_value += df.query('Site == "'+site+'" & Unit == "'+unit+'" & Scheme == "American Express"')['TVal'].sum()
                sum_transaction_charge += (df.query('Site == "'+site+'" & Unit == "'+unit+'" & Scheme == "American Express"')['Service Charge Value Corrected'].sum())
                sum_transaction_charge += (df.query('Site == "'+site+'" & Unit == "'+unit+'" & Scheme == "American Express"')['Standard Charge Value'].sum())
                sum_standard_charge += df.query('Site == "'+site+'" & Unit == "'+unit+'" & Scheme == "American Express"')['Service Charge Value'].sum()
                sum_service_charge = (sum_standard_charge/100)*20
                sum_transaction_value = float("{:.2f}".format(sum_transaction_value))
                sum_transaction_charge = float("{:.2f}".format(sum_transaction_charge))
                sum_standard_charge = float("{:.2f}".format(sum_standard_charge))
                sum_service_charge = float("{:.2f}".format(sum_service_charge))
                totals.append(Unit(site, unit, "Amex", sum_transaction_value, sum_transaction_charge, sum_standard_charge, sum_service_charge))

    
    ws = template.active
    read_range = ws['F10':'F146']
    write_range = ws['C10':'C146']

    for record in totals:
        if 'Costa' in record.unit:
            record.unit = record.unit.replace('Coffee', '').strip()
        if 'WHS' in record.unit:
            record.unit = record.unit.replace('WHSmiths', 'WHSmith')
        if 'North' in record.site:
            record.site = record.site.replace('(North)', 'Nth').strip()
        if 'South' in record.site:
            record.site = record.site.replace('(South)', 'Sth').strip()
            
        
    for record in totals:
        print(record.site, record.unit, record.scheme, record.sum_transaction_value, record.sum_transaction_charge, record.sum_standard_charge, record.sum_service_charge)

    i=0
    for cell in read_range:
        for record in totals:
            # print(cell[0].value, record.site + " " + record.unit)
            if cell[0].value == (record.site + " " + record.unit + " " + record.scheme).strip(): 
                write_range[i][0].value = record.sum_transaction_value
                write_range[i+1][0].value = record.sum_transaction_charge
                write_range[i+2][0].value = record.sum_standard_charge
                write_range[i+3][0].value = record.sum_service_charge
        i += 1

    DataFrame.to_csv(df, 'output/output.csv', index=False)
    template.save('output/output.xlsx')